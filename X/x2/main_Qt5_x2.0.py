# MIT License
#
# Copyright (c) [2026] [DZMMC]
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

import sys
import random
import csv
import logging
import os
import time
from datetime import datetime
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QHBoxLayout, QPushButton, QLabel, QLineEdit,
                             QListWidget, QListWidgetItem, QMessageBox, QSplitter,
                             QFileDialog, QTextBrowser, QDialog, QDialogButtonBox,
                             QSizePolicy)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QMutex, QMutexLocker

# 导入openpyxl库（用于XLSX导出）
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
except ImportError:
    QMessageBox.critical(None, "依赖缺失", "请先安装openpyxl库：pip install openpyxl")
    sys.exit(1)

# 尝试导入xlwt库（用于XLS导出）
try:
    import xlwt
except ImportError:
    xlwt = None


# 自定义日志Handler，用于将日志同时输出到文件和界面
class QTextBrowserHandler(logging.Handler):
    def __init__(self, text_browser):
        super().__init__()
        self.text_browser = text_browser
        # 设置日志格式
        formatter = logging.Formatter(
            '%(asctime)s - %(levelname)s - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        self.setFormatter(formatter)

    def emit(self, record):
        # 格式化日志信息
        msg = self.format(record)
        # 在界面中显示日志
        self.text_browser.append(msg)
        # 自动滚动到最新日志
        self.text_browser.verticalScrollBar().setValue(
            self.text_browser.verticalScrollBar().maximum()
        )


# 日志查看窗口
class LogViewerDialog(QDialog):
    def __init__(self, log_file_path, parent=None):
        super().__init__(parent)
        self.log_file_path = log_file_path
        self.setWindowTitle("查看日志")
        self.setGeometry(150, 150, 800, 600)
        self.setModal(True)

        # 创建布局
        layout = QVBoxLayout(self)

        # 日志显示区域
        self.log_text = QTextBrowser()
        self.log_text.setReadOnly(True)
        layout.addWidget(self.log_text)

        # 按钮区域
        button_box = QDialogButtonBox()
        self.refresh_btn = QPushButton("刷新日志")
        self.refresh_btn.clicked.connect(self.load_logs)
        button_box.addButton(self.refresh_btn, QDialogButtonBox.ActionRole)
        button_box.addButton(QDialogButtonBox.Close)
        button_box.rejected.connect(self.close)
        layout.addWidget(button_box)

        # 加载日志内容
        self.load_logs()

    def load_logs(self):
        """加载日志文件内容"""
        try:
            if os.path.exists(self.log_file_path):
                with open(self.log_file_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                    self.log_text.setText(content)
            else:
                self.log_text.setText("日志文件不存在")
        except Exception as e:
            QMessageBox.warning(self, "错误", f"加载日志失败:\n{str(e)}")
            logging.error(f"加载日志失败: {str(e)}")


# 后台工作线程：处理抽号和分组任务
class WorkerThread(QThread):
    # 定义信号
    task_finished = pyqtSignal(object)  # 任务完成信号，返回结果
    error_occurred = pyqtSignal(str)    # 错误信号，返回错误信息

    def __init__(self, task_type, *args, **kwargs):
        super().__init__()
        self.task_type = task_type
        self.args = args
        self.kwargs = kwargs
        self.mutex = QMutex()

    def run(self):
        """线程执行入口"""
        try:
            if self.task_type == "draw":
                # 抽号任务
                all_numbers = self.args[0]
                with QMutexLocker(self.mutex):
                    if not all_numbers:
                        self.error_occurred.emit("号码池已空")
                        return
                    number = random.choice(all_numbers)
                    all_numbers.remove(number)
                self.task_finished.emit({"type": "draw", "number": number})
            elif self.task_type == "group":
                # 分组任务
                drawn_numbers = self.args[0]
                group_size = self.args[1]
                numbers_copy = drawn_numbers.copy()
                random.shuffle(numbers_copy)
                groups = []
                for i in range(0, len(numbers_copy), group_size):
                    group_members = numbers_copy[i:i + group_size]
                    groups.append(group_members)
                self.task_finished.emit({"type": "group", "groups": groups})
        except Exception as e:
            self.error_occurred.emit(f"任务执行失败: {str(e)}")


class DrawAndGroupSystem(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("抽号与分组系统")
        self.setGeometry(100, 100, 1000, 600)

        # 初始化数据
        self.all_numbers = []
        self.drawn_numbers = []
        self.groups = []
        self.worker_thread = None

        # 初始化日志相关
        self.log_dir = "logs"
        self.export_dir = "exports"
        self.log_file_path = os.path.join(self.log_dir, "system.log")
        self._init_directories()
        self._init_logging()

        # 创建中心部件和主布局
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        # 顶部功能按钮区域
        top_layout = QHBoxLayout()

        # 查看日志按钮
        self.view_log_btn = QPushButton("查看日志")
        self.view_log_btn.clicked.connect(self.open_log_viewer)
        top_layout.addWidget(self.view_log_btn)

        # 清理缓存按钮
        self.clear_cache_btn = QPushButton("清理缓存/日志")
        self.clear_cache_btn.clicked.connect(self.clear_cache_and_logs)
        top_layout.addWidget(self.clear_cache_btn)

        top_layout.addStretch()
        main_layout.addLayout(top_layout)

        # 日志显示区域
        self.log_display = QTextBrowser()
        self.log_display.setMaximumHeight(100)
        self.log_display.setReadOnly(True)
        main_layout.addWidget(self.log_display)

        # 将自定义Handler添加到logger
        self.log_handler = QTextBrowserHandler(self.log_display)
        logging.getLogger().addHandler(self.log_handler)

        # 创建左右分割面板
        splitter = QSplitter(Qt.Horizontal)
        main_layout.addWidget(splitter)

        # ---------------------- 左侧：抽号区域 ----------------------
        draw_widget = QWidget()
        draw_layout = QVBoxLayout(draw_widget)

        # 抽号标题
        draw_title = QLabel("抽号系统")
        draw_title.setAlignment(Qt.AlignCenter)
        draw_title.setStyleSheet("font-size: 20px; font-weight: bold; margin-bottom: 15px;")
        draw_layout.addWidget(draw_title)

        # 范围输入区域
        range_layout = QHBoxLayout()
        range_layout.addWidget(QLabel("起始号码:"))
        self.start_input = QLineEdit()
        self.start_input.setText("1")
        range_layout.addWidget(self.start_input)

        range_layout.addWidget(QLabel("结束号码:"))
        self.end_input = QLineEdit()
        self.end_input.setText("100")
        range_layout.addWidget(self.end_input)

        self.init_button = QPushButton("初始化号码池")
        self.init_button.clicked.connect(self.init_number_pool)
        range_layout.addWidget(self.init_button)

        draw_layout.addLayout(range_layout)

        # 抽号控制区域
        draw_control_layout = QHBoxLayout()
        self.draw_button = QPushButton("抽取号码")
        self.draw_button.clicked.connect(self.draw_number)
        self.draw_button.setEnabled(False)
        draw_control_layout.addWidget(self.draw_button)

        self.reset_button = QPushButton("重置抽号")
        self.reset_button.clicked.connect(self.reset_draw)
        self.reset_button.setEnabled(False)
        draw_control_layout.addWidget(self.reset_button)

        # 新增导出按钮
        self.export_draw_button = QPushButton("导出抽号结果")
        self.export_draw_button.clicked.connect(self.export_draw_results)
        self.export_draw_button.setEnabled(False)
        draw_control_layout.addWidget(self.export_draw_button)

        draw_layout.addLayout(draw_control_layout)

        # 当前抽到的号码
        self.current_number_label = QLabel("当前抽到的号码: ")
        self.current_number_label.setAlignment(Qt.AlignCenter)
        self.current_number_label.setStyleSheet(
            "font-size: 18px; margin: 20px 0; padding: 10px; background-color: #f0f0f0; border-radius: 5px;")
        draw_layout.addWidget(self.current_number_label)

        # 已抽号码列表
        drawn_label = QLabel("已抽号码:")
        draw_layout.addWidget(drawn_label)

        self.drawn_list = QListWidget()
        draw_layout.addWidget(self.drawn_list)

        # 添加到分割面板
        splitter.addWidget(draw_widget)

        # ---------------------- 右侧：分组区域 ----------------------
        group_widget = QWidget()
        group_layout = QVBoxLayout(group_widget)

        # 分组标题
        group_title = QLabel("分组系统")
        group_title.setAlignment(Qt.AlignCenter)
        group_title.setStyleSheet("font-size: 20px; font-weight: bold; margin-bottom: 15px;")
        group_layout.addWidget(group_title)

        # 分组设置区域
        group_setting_layout = QHBoxLayout()
        group_setting_layout.addWidget(QLabel("每组人数:"))
        self.group_size_input = QLineEdit()
        self.group_size_input.setText("5")
        group_setting_layout.addWidget(self.group_size_input)

        self.group_button = QPushButton("进行分组")
        self.group_button.clicked.connect(self.create_groups)
        self.group_button.setEnabled(False)
        group_setting_layout.addWidget(self.group_button)

        # 新增导出分组按钮
        self.export_group_button = QPushButton("导出分组结果")
        self.export_group_button.clicked.connect(self.export_group_results)
        self.export_group_button.setEnabled(False)
        group_setting_layout.addWidget(self.export_group_button)

        group_layout.addLayout(group_setting_layout)

        # 分组结果区域
        result_label = QLabel("分组结果:")
        group_layout.addWidget(result_label)

        self.group_list = QListWidget()
        group_layout.addWidget(self.group_list)

        # 添加到分割面板
        splitter.addWidget(group_widget)

        # 设置分割面板初始比例
        splitter.setSizes([500, 500])

        # 记录程序启动日志
        logging.info("抽号与分组系统启动成功")

    def _init_directories(self):
        """初始化必要目录"""
        for dir_path in [self.log_dir, self.export_dir]:
            if not os.path.exists(dir_path):
                os.makedirs(dir_path)
                logging.info(f"创建目录: {dir_path}")

    def _init_logging(self):
        """初始化日志系统"""
        # 配置日志记录器
        logger = logging.getLogger()
        logger.setLevel(logging.INFO)

        # 移除已有的Handler，避免重复输出
        if logger.handlers:
            for handler in logger.handlers:
                if not isinstance(handler, QTextBrowserHandler):
                    logger.removeHandler(handler)

        # 创建文件Handler，将日志写入文件
        file_handler = logging.FileHandler(
            self.log_file_path,
            mode='a',
            encoding='utf-8'
        )
        file_handler.setLevel(logging.INFO)

        # 创建控制台Handler，将日志输出到控制台
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)

        # 设置日志格式
        formatter = logging.Formatter(
            '%(asctime)s - %(levelname)s - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        file_handler.setFormatter(formatter)
        console_handler.setFormatter(formatter)

        # 将Handler添加到logger
        logger.addHandler(file_handler)
        logger.addHandler(console_handler)

    def open_log_viewer(self):
        """打开日志查看窗口"""
        try:
            log_dialog = LogViewerDialog(self.log_file_path, self)
            log_dialog.exec_()
        except Exception as e:
            QMessageBox.critical(self, "错误", f"打开日志查看器失败:\n{str(e)}")
            logging.error(f"打开日志查看器失败: {str(e)}")

    def clear_cache_and_logs(self):
        """清理缓存和日志文件"""
        reply = QMessageBox.question(
            self,
            "确认清理",
            "确定要清理所有缓存和日志文件吗？\n此操作不可恢复！",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            try:
                # 关闭当前日志文件
                logger = logging.getLogger()
                for handler in logger.handlers:
                    if isinstance(handler, logging.FileHandler):
                        handler.close()
                        logger.removeHandler(handler)

                # 删除日志文件
                if os.path.exists(self.log_file_path):
                    os.remove(self.log_file_path)
                    logging.info("日志文件已清理")

                # 删除导出文件
                for file_name in os.listdir(self.export_dir):
                    file_path = os.path.join(self.export_dir, file_name)
                    if os.path.isfile(file_path):
                        os.remove(file_path)
                        logging.info(f"删除导出文件: {file_path}")

                # 重新初始化日志系统
                self._init_logging()
                self.log_display.clear()

                QMessageBox.information(self, "成功", "缓存和日志文件已清理完成")
                logging.info("缓存清理操作完成")
            except Exception as e:
                QMessageBox.critical(self, "错误", f"清理缓存失败:\n{str(e)}")
                logging.error(f"清理缓存失败: {str(e)}")

    def init_number_pool(self):
        """初始化号码池"""
        try:
            start = int(self.start_input.text())
            end = int(self.end_input.text())

            if start >= end:
                QMessageBox.warning(self, "输入错误", "起始号码必须小于结束号码")
                logging.warning(f"初始化号码池失败: 起始号码{start}不小于结束号码{end}")
                return

            self.all_numbers = list(range(start, end + 1))
            self.drawn_numbers = []
            self.drawn_list.clear()
            self.current_number_label.setText("当前抽到的号码: ")
            self.group_list.clear()

            self.draw_button.setEnabled(True)
            self.reset_button.setEnabled(True)
            self.group_button.setEnabled(False)
            self.export_draw_button.setEnabled(False)
            self.export_group_button.setEnabled(False)

            success_msg = f"号码池已初始化，共{end - start + 1}个号码"
            QMessageBox.information(self, "成功", success_msg)
            logging.info(success_msg)

        except ValueError:
            QMessageBox.warning(self, "输入错误", "请输入有效的整数")
            logging.warning("初始化号码池失败: 输入不是有效的整数")

    def draw_number(self):
        """抽取一个号码（后台线程执行）"""
        if self.worker_thread and self.worker_thread.isRunning():
            return

        self.draw_button.setEnabled(False)
        self.worker_thread = WorkerThread("draw", self.all_numbers)
        self.worker_thread.task_finished.connect(self.on_draw_finished)
        self.worker_thread.error_occurred.connect(self.on_worker_error)
        self.worker_thread.start()

    def on_draw_finished(self, result):
        """抽号任务完成回调"""
        try:
            if result["type"] == "draw":
                number = result["number"]
                self.drawn_numbers.append(number)
                # 更新显示
                self.current_number_label.setText(f"当前抽到的号码: {number}")
                self.drawn_list.addItem(QListWidgetItem(str(number)))
                # 启用分组和导出按钮
                self.group_button.setEnabled(True)
                self.export_draw_button.setEnabled(True)
                self.draw_button.setEnabled(True)

                logging.info(f"抽取号码: {number}，剩余号码: {len(self.all_numbers)}个")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"处理抽号结果失败:\n{str(e)}")
            logging.error(f"处理抽号结果失败: {str(e)}")
            self.draw_button.setEnabled(True)

    def on_worker_error(self, error_msg):
        """线程错误回调"""
        QMessageBox.warning(self, "操作失败", error_msg)
        logging.error(f"后台任务失败: {error_msg}")
        self.draw_button.setEnabled(True)
        self.group_button.setEnabled(True)

    def reset_draw(self):
        """重置抽号"""
        self.all_numbers = []
        self.drawn_numbers = []
        self.groups = []
        self.drawn_list.clear()
        self.current_number_label.setText("当前抽到的号码: ")
        self.group_list.clear()
        self.draw_button.setEnabled(False)
        self.reset_button.setEnabled(False)
        self.group_button.setEnabled(False)
        self.export_draw_button.setEnabled(False)
        self.export_group_button.setEnabled(False)

        logging.info("抽号系统已重置")

    def create_groups(self):
        """进行分组（后台线程执行）"""
        try:
            group_size = int(self.group_size_input.text())

            if group_size <= 0:
                QMessageBox.warning(self, "输入错误", "每组人数必须大于0")
                logging.warning(f"分组失败: 每组人数{group_size}必须大于0")
                return

            if len(self.drawn_numbers) < group_size:
                QMessageBox.warning(self, "输入错误", "每组人数不能大于已抽号码数量")
                logging.warning(f"分组失败: 每组人数{group_size}大于已抽号码数量{len(self.drawn_numbers)}")
                return

            if self.worker_thread and self.worker_thread.isRunning():
                return

            self.group_button.setEnabled(False)
            self.worker_thread = WorkerThread("group", self.drawn_numbers, group_size)
            self.worker_thread.task_finished.connect(self.on_group_finished)
            self.worker_thread.error_occurred.connect(self.on_worker_error)
            self.worker_thread.start()

        except ValueError:
            QMessageBox.warning(self, "输入错误", "请输入有效的整数")
            logging.warning("分组失败: 输入不是有效的整数")

    def on_group_finished(self, result):
        """分组任务完成回调"""
        try:
            if result["type"] == "group":
                self.groups = result["groups"]
                self.group_list.clear()
                for i, group in enumerate(self.groups, 1):
                    group_item = QListWidgetItem(f"第{i}组: {', '.join(map(str, group))}")
                    self.group_list.addItem(group_item)

                # 启用导出分组按钮
                self.export_group_button.setEnabled(True)
                self.group_button.setEnabled(True)

                success_msg = f"已成功创建{len(self.groups)}个分组，每组{len(self.groups[0])}人"
                logging.info(success_msg)

        except Exception as e:
            QMessageBox.critical(self, "错误", f"处理分组结果失败:\n{str(e)}")
            logging.error(f"处理分组结果失败: {str(e)}")
            self.group_button.setEnabled(True)

    def export_draw_results(self):
        """导出抽号结果为CSV/XLSX"""
        if not self.drawn_numbers:
            QMessageBox.warning(self, "提示", "没有可导出的抽号数据")
            return

        # 选择导出格式和路径
        file_filters = "ExcelX文件 (*.xlsx);;CSV文件 (*.csv)"
        if xlwt is not None:
            file_filters += ";;Excel文件 (*.xls)"
        
        file_path, file_type = QFileDialog.getSaveFileName(
            self, "保存抽号结果", os.path.join(self.export_dir, "抽号结果"),
            file_filters
        )

        if not file_path:
            return  # 用户取消保存

        try:
            if file_type == "Excel文件 (*.xls)" and xlwt is not None:
                self._export_draw_to_xls(file_path)
            elif file_type == "CSV文件 (*.csv)":
                self._export_draw_to_csv(file_path)
            else:
                # 默认导出为XLSX
                self._export_draw_to_xlsx(file_path)

            success_msg = f"抽号结果已成功导出到:\n{file_path}"
            QMessageBox.information(self, "成功", success_msg)
            logging.info(success_msg)

        except PermissionError:
            QMessageBox.critical(self, "错误", "没有文件写入权限，请检查文件是否被占用或尝试其他路径")
            logging.error(f"导出失败: 没有权限写入文件 {file_path}")
        except Exception as e:
            error_msg = f"导出失败:\n{str(e)}"
            QMessageBox.critical(self, "错误", error_msg)
            logging.error(f"抽号结果导出失败: {str(e)}")

    def export_group_results(self):
        """导出分组结果为CSV/XLSX"""
        if not self.groups:
            QMessageBox.warning(self, "提示", "没有可导出的分组数据")
            return

        # 选择导出格式和路径
        file_filters = "ExcelX文件 (*.xlsx);;CSV文件 (*.csv)"
        if xlwt is not None:
            file_filters += ";;Excel文件 (*.xls)"
        
        file_path, file_type = QFileDialog.getSaveFileName(
            self, "保存分组结果", os.path.join(self.export_dir, "分组结果"),
            file_filters
        )

        if not file_path:
            return  # 用户取消保存

        try:
            if file_type == "Excel文件 (*.xls)" and xlwt is not None:
                self._export_groups_to_xls(file_path)
            elif file_type == "CSV文件 (*.csv)":
                self._export_groups_to_csv(file_path)
            else:
                # 默认导出为XLSX
                self._export_groups_to_xlsx(file_path)

            success_msg = f"分组结果已成功导出到:\n{file_path}"
            QMessageBox.information(self, "成功", success_msg)
            logging.info(success_msg)

        except PermissionError:
            QMessageBox.critical(self, "错误", "没有文件写入权限，请检查文件是否被占用或尝试其他路径")
            logging.error(f"导出失败: 没有权限写入文件 {file_path}")
        except Exception as e:
            error_msg = f"导出失败:\n{str(e)}"
            QMessageBox.critical(self, "错误", error_msg)
            logging.error(f"分组结果导出失败: {str(e)}")

    def _export_draw_to_csv(self, file_path):
        """内部方法：将抽号结果导出为CSV"""
        with open(file_path, mode='w', newline='', encoding='utf-8-sig') as file:
            writer = csv.writer(file)
            # 写入表头
            writer.writerow(["抽号顺序", "号码"])
            # 写入数据
            for idx, number in enumerate(self.drawn_numbers, 1):
                writer.writerow([idx, number])

    def _export_draw_to_xlsx(self, file_path):
        """内部方法：将抽号结果导出为XLSX"""
        wb = Workbook()
        ws = wb.active
        ws.title = "抽号结果"

        # 设置表头样式
        header_font = Font(bold=True, size=12)
        header_alignment = Alignment(horizontal='center')

        # 写入表头
        headers = ["抽号顺序", "号码"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = header_alignment

        # 写入数据
        for row, number in enumerate(self.drawn_numbers, 2):
            ws.cell(row=row, column=1, value=row - 1)  # 抽号顺序
            ws.cell(row=row, column=2, value=number)  # 号码

        # 调整列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12

        # 保存文件
        wb.save(file_path)

    def _export_draw_to_xls(self, file_path):
        """内部方法：将抽号结果导出为XLS"""
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet("抽号结果")

        # 设置表头样式
        header_style = xlwt.XFStyle()
        font = xlwt.Font()
        font.bold = True
        font.height = 240  # 12pt
        header_style.font = font
        alignment = xlwt.Alignment()
        alignment.horz = xlwt.Alignment.HORZ_CENTER
        header_style.alignment = alignment

        # 写入表头
        headers = ["抽号顺序", "号码"]
        for col, header in enumerate(headers):
            ws.write(0, col, header, header_style)

        # 写入数据
        for row, number in enumerate(self.drawn_numbers, 1):
            ws.write(row, 0, row)  # 抽号顺序
            ws.write(row, 1, number)  # 号码

        # 调整列宽
        ws.col(0).width = 256 * 12
        ws.col(1).width = 256 * 12

        # 保存文件
        wb.save(file_path)

    def _export_groups_to_csv(self, file_path):
        """内部方法：将分组结果导出为CSV"""
        with open(file_path, mode='w', newline='', encoding='utf-8-sig') as file:
            writer = csv.writer(file)
            # 写入表头
            writer.writerow(["组号", "成员编号"])
            # 写入数据
            for group_idx, group in enumerate(self.groups, 1):
                for member in group:
                    writer.writerow([group_idx, member])

    def _export_groups_to_xlsx(self, file_path):
        """内部方法：将分组结果导出为XLSX"""
        wb = Workbook()
        ws = wb.active
        ws.title = "分组结果"

        # 设置表头样式
        header_font = Font(bold=True, size=12)
        header_alignment = Alignment(horizontal='center')

        # 写入表头
        headers = ["组号", "成员编号"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = header_alignment

        # 写入数据
        row = 2
        for group_idx, group in enumerate(self.groups, 1):
            for member in group:
                ws.cell(row=row, column=1, value=group_idx)
                ws.cell(row=row, column=2, value=member)
                row += 1

        # 调整列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12

        # 保存文件
        wb.save(file_path)

    def _export_groups_to_xls(self, file_path):
        """内部方法：将分组结果导出为XLS"""
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet("分组结果")

        # 设置表头样式
        header_style = xlwt.XFStyle()
        font = xlwt.Font()
        font.bold = True
        font.height = 240  # 12pt
        header_style.font = font
        alignment = xlwt.Alignment()
        alignment.horz = xlwt.Alignment.HORZ_CENTER
        header_style.alignment = alignment

        # 写入表头
        headers = ["组号", "成员编号"]
        for col, header in enumerate(headers):
            ws.write(0, col, header, header_style)

        # 写入数据
        row = 1
        for group_idx, group in enumerate(self.groups, 1):
            for member in group:
                ws.write(row, 0, group_idx)
                ws.write(row, 1, member)
                row += 1

        # 调整列宽
        ws.col(0).width = 256 * 12
        ws.col(1).width = 256 * 12

        # 保存文件
        wb.save(file_path)


# 主程序
if __name__ == "__main__":
    app = QApplication(sys.argv)
    main_window = DrawAndGroupSystem()
    main_window.show()
    sys.exit(app.exec_())